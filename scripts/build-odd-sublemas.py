#!/usr/bin/env python3
"""
build-odd-sublemas.py — inyecta el nivel SUBLEMA en la contienda ODD del catálogo de internas.

El acordeón (OpcionAccordion) y el gate de escaleras ya soportan lema → sublema → hoja: nacionales
lo usa (Story 10.8) con id de nodo `{lemaId}-sl-{slug(sublema)}` y `grupoId`/`sublemaId` en la opción.
El ODD de internas declaraba `["lema","hoja"]` porque el desglose-de-votos NO trae la columna SUBLEMA;
sí la trae la "Integración hojas de votación" de la Corte (hoja → sublema, por TipoHoja=ODD).

Este post-step (espeja build-hoja-local.py: Python sobre los shards ya generados) PARCHEA en sitio el
`catalogo.json` de cada depto:
  - agrega nodos `sublema` (parentId = lemaId),
  - pone `grupoId`+`sublemaId` en cada opción ODD agrupada,
  - sube `niveles` a ["lema","sublema","hoja"] (o deja ["lema","hoja"] + degradado:true si ese depto
    no tiene ningún sublema con dato).
Preserva los lemaId "bare" del sweep (parchea, no regenera) y no toca los shards de votos. Idempotente.

REQUIERE que ESCALERAS declare (internas, odd) = ['lema','sublema','hoja'] (granularidad.ts).

Fuente de integración (auto-detección por elección):
  internas-2024 → data/raw/electoral/internas-2024/integracion-hojas-de-votacion.xlsx (hoja 'Datos')
  internas-2019 → data/raw/electoral/internas-2019/integraci-n-de-las-hojas-de-votaci-n.csv
(el CSV de 2024 viene vacío del origen → se usa el XLSX).

Uso:
  python scripts/build-odd-sublemas.py internas-2024
  python scripts/build-odd-sublemas.py internas-2019
"""
import csv, json, os, re, glob, sys, unicodedata, argparse
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# folder de depto → código de 2 letras de la Corte (inverso del CODE_DIR de build-hoja-local.py)
DEPT_CODE = {"artigas":"AR","canelones":"CA","cerro_largo":"CL","colonia":"CO","durazno":"DU",
    "florida":"FD","flores":"FS","lavalleja":"LA","maldonado":"MA","montevideo":"MO","paysandu":"PA",
    "rio_negro":"RN","rocha":"RO","rivera":"RV","salto":"SA","san_jose":"SJ","soriano":"SO",
    "tacuarembo":"TA","treinta_y_tres":"TT"}

SIN_SUBLEMA = {"SIN SUBLEMA", "NO APLICA", "NO APLICA.", ""}


def norm(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper().strip()


def norm_strip(s):
    """norm + saca el prefijo 'PARTIDO ' → une 'PARTIDO NACIONAL' ↔ 'Partido Nacional' ↔ 'Nacional'."""
    return re.sub(r"^PARTIDO\s+", "", norm(s))


def slug(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def integracion_source(eleccion):
    base = os.path.join(ROOT, "data/raw/electoral", eleccion)
    # candidatos CSV no vacíos
    for name in os.listdir(base):
        low = name.lower()
        if "integ" in low and "hoja" in low and low.endswith(".csv"):
            p = os.path.join(base, name)
            if os.path.getsize(p) > 0:
                return ("csv", p)
    for name in os.listdir(base):
        low = name.lower()
        if "integ" in low and "hoja" in low and low.endswith(".xlsx"):
            return ("xlsx", os.path.join(base, name))
    raise SystemExit(f"  ✗ sin integración de hojas en {base}")


def read_integracion_rows(kind, path):
    """Itera filas como dict con claves Departamento, TipoHoja, PartidoPolitico, Numero, Sublema."""
    if kind == "csv":
        enc = "utf-8-sig"
        try: open(path, encoding=enc).readline()
        except Exception: enc = "latin-1"
        with open(path, encoding=enc, errors="replace", newline="") as f:
            for row in csv.DictReader(f):
                yield row
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Datos"] if "Datos" in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(h) if h is not None else "" for h in next(it)]
        idx = {h: i for i, h in enumerate(hdr)}
        for row in it:
            yield {h: row[idx[h]] for h in hdr}


def build_sublema_map(eleccion):
    """Devuelve (submap, present):
       submap[(deptCode, normStripPartido, hojaStr)] = sublemaRaw — solo TipoHoja=ODD con sublema real.
       present = set de (deptCode, normStripPartido, hojaStr) de TODA hoja ODD (incl. 'Sin sublema'),
                 para distinguir 'directa por Sin sublema' (OK) de 'ausente de integración' (bug de join)."""
    kind, path = integracion_source(eleccion)
    print(f"  fuente integración: [{kind}] {os.path.relpath(path, ROOT)}")
    m = {}                       # (dep, partKey, hoja) -> sublema
    present = set()
    conflicts = 0
    for row in read_integracion_rows(kind, path):
        if str(row.get("TipoHoja", "")).strip().upper() != "ODD":
            continue
        dep = str(row.get("Departamento", "") or "").strip().upper()
        part = norm_strip(str(row.get("PartidoPolitico", "") or ""))
        hoja = str(row.get("Numero", "") or "").strip()
        if hoja.endswith(".0"):  # XLSX numérico → "969.0"
            hoja = hoja[:-2]
        if not (dep and part and hoja):
            continue
        key = (dep, part, hoja)
        present.add(key)
        sub = str(row.get("Sublema", "") or "").strip()
        if norm(sub) in SIN_SUBLEMA:
            continue
        if key in m and m[key] != sub:
            conflicts += 1
        m[key] = sub
    if conflicts:
        print(f"  ⚠ {conflicts} (dep,partido,hoja) con sublema ambiguo (gana el último)")
    return m, present


def patch_catalogo(path, dep_code, submap, present):
    doc = json.load(open(path, encoding="utf-8"))
    odd = next((c for c in doc.get("contiendas", []) if c.get("contienda") == "odd"), None)
    if not odd:
        return None
    # lemaId → clave de partido normalizada (desde la etiqueta del nodo lema)
    lema_key = {n["id"]: norm_strip(n.get("etiqueta", "")) for n in odd["nodos"] if n.get("nivel") == "lema"}
    # idempotencia: sacar nodos sublema previos y limpiar grupoId/sublemaId
    odd["nodos"] = [n for n in odd["nodos"] if n.get("nivel") != "sublema"]
    sub_nodes = {}                   # subId -> nodo
    n_group = n_direct = n_unmatched = 0
    for op in odd.get("opciones", []):
        op.pop("grupoId", None); op.pop("sublemaId", None)
        lemaId = op.get("lemaId", "")
        hoja = str(op.get("hoja", ""))
        part = lema_key.get(lemaId, "")
        sub = submap.get((dep_code, part, hoja))
        if not sub:
            # cuelgan directo del lema. 'vl' y 'Sin sublema' son esperados; ausente de integración = bug de join.
            if hoja and hoja != "vl" and (dep_code, part, hoja) not in present:
                n_unmatched += 1
            n_direct += 1
            continue
        subId = f"{lemaId}-sl-{slug(sub)}"
        if subId not in sub_nodes:
            sub_nodes[subId] = {"id": subId, "nivel": "sublema", "etiqueta": sub, "parentId": lemaId}
        op["sublemaId"] = subId
        op["grupoId"] = subId
        n_group += 1
    odd["nodos"].extend(sub_nodes.values())
    if sub_nodes:
        odd["niveles"] = ["lema", "sublema", "hoja"]
        odd.pop("degradado", None)
    else:
        odd["niveles"] = ["lema", "hoja"]
        odd["degradado"] = True      # gate: niveles más cortos que la escalera requieren degradado
    json.dump(doc, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    with open(path, "a", encoding="utf-8") as f:
        f.write("\n")
    return {"sublemas": len(sub_nodes), "agrupadas": n_group, "directas": n_direct, "sin_match": n_unmatched}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("eleccion", help="internas-2024 | internas-2019")
    a = ap.parse_args()

    print(f"=== Inyección de SUBLEMA en ODD — {a.eleccion} ===")
    submap, present = build_sublema_map(a.eleccion)
    print(f"  mapa hoja→sublema (ODD): {len(submap)} entradas · {len(present)} hojas ODD en integración")

    cats = sorted(glob.glob(os.path.join(ROOT, f"public/data/{a.eleccion}/*/catalogo.json")))
    tot_sub = tot_match = tot_unmatch = 0
    for p in cats:
        dep = os.path.basename(os.path.dirname(p))
        code = DEPT_CODE.get(dep)
        if not code:
            print(f"  · {dep}: sin código de depto — skip"); continue
        r = patch_catalogo(p, code, submap, present)
        if r is None:
            print(f"  · {dep}: sin contienda odd — skip"); continue
        tot_sub += r["sublemas"]; tot_match += r["agrupadas"]; tot_unmatch += r["sin_match"]
        flag = "" if r["sublemas"] else "  (degradado: sin sublemas)"
        warn = f" ⚠{r['sin_match']} hojas sin match" if r["sin_match"] else ""
        print(f"  ✓ {dep:16s} {r['sublemas']:3d} sublemas · {r['agrupadas']:4d} hojas agrupadas · {r['directas']:4d} directas{warn}{flag}")
    print(f"Listo. {tot_sub} nodos sublema · {tot_match} hojas agrupadas · {tot_unmatch} sin match.")


if __name__ == "__main__":
    main()
