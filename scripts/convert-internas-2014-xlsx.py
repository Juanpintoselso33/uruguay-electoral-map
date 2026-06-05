#!/usr/bin/env python3
"""
convert-internas-2014-xlsx.py — Convierte los 20 XLSX crudos de internas-2014 (Corte Electoral,
directo, NO datos abiertos) a un `desglose-de-votos.csv` UTF-8 con el MISMO esquema que consume
el pipeline (igual que internas-2019/2024):

    TIPO_REGISTRO,DEPARTAMENTO,CRV,SERIES,LEMA,DESCRIPCIÓN_1,DESCRIPCIÓN_2,CANTIDAD_VOTOS

Entrada (8 filas de título, tabla desde fila 9):
  - odn/*.xlsx (9 cols): ACTO,CONVOCATORIA,DEPTO,CIRCUITO,SERIES,ESCRUTINIO,PRECANDIDATO,HOJA,CNT_VOTOS
      → TIPO_REGISTRO=HOJA_ODN, DESCRIPCIÓN_1=PRECANDIDATO, DESCRIPCIÓN_2=HOJA
  - odd/*.xlsx (8 cols, SIN precandidato): ...,HOJA,CNT_VOTOS
      → TIPO_REGISTRO=HOJA_ODD, DESCRIPCIÓN_1='No aplica', DESCRIPCIÓN_2=HOJA
  - CASO ESPECIAL: 'DE LA CONCERTACION ODN.xlsx' viene en formato ANCHO (una fila por CRV con
      columnas HABILITADO/EN_BLANCO/.../JOSE LUIS VERA). Tiene un único precandidato (José Luis
      Vera) y NO trae nº de hoja → se emite como HOJA_ODN con DESCRIPCIÓN_1='JOSE LUIS VERA',
      DESCRIPCIÓN_2='No aplica' (→ el runner lo trata como voto-al-lema, hoja=VL). Las columnas
      EN_BLANCO/ANULADOS/OBSERVADOS de ese export son del CIRCUITO completo (todos los partidos),
      NO de este lema → se IGNORAN (no son noPartidarios del partido).

LEMA = CONVOCATORIA sin el sufijo "_-_O.D.x", underscores → espacios, encoding reparado.
DEPARTAMENTO = código de 2 letras (Montevideo→MO, …), match por nombre normalizado sin acentos
  (el crudo viene con acentos comidos: "Paysand", "Ro Negro", "San Jos", "Tacuaremb").

Uso: python scripts/convert-internas-2014-xlsx.py
"""
import csv, glob, os, re, unicodedata
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data/raw/electoral/internas-2014")
OUT = os.path.join(RAW, "desglose-de-votos.csv")

# nombre normalizado (sin acentos, mayúsculas) → código de 2 letras
NAME2CODE = {
    "ARTIGAS": "AR", "CANELONES": "CA", "CERRO LARGO": "CL", "COLONIA": "CO",
    "DURAZNO": "DU", "FLORES": "FS", "FLORIDA": "FD", "LAVALLEJA": "LA",
    "MALDONADO": "MA", "MONTEVIDEO": "MO", "PAYSANDU": "PA", "RIO NEGRO": "RN",
    "RIVERA": "RV", "ROCHA": "RO", "SALTO": "SA", "SAN JOSE": "SJ",
    "SORIANO": "SO", "TACUAREMBO": "TA", "TREINTA Y TRES": "TT",
}
# alias para nombres con acentos comidos por el export
NAME_ALIAS = {
    "PAYSAND": "PAYSANDU", "RO NEGRO": "RIO NEGRO", "SAN JOS": "SAN JOSE",
    "TACUAREMB": "TACUAREMBO",
}


def nfd(s):
    s = unicodedata.normalize("NFD", s or "")
    return "".join(c for c in s if unicodedata.category(c) != "Mn").upper().strip()


def dep_code(raw):
    n = nfd(raw)
    n = NAME_ALIAS.get(n, n)
    code = NAME2CODE.get(n)
    if not code:
        raise ValueError(f"DEPTO desconocido: {raw!r} (nfd={n!r})")
    return code


def lema_from_convocatoria(conv):
    """'Frente_Amplio_-_O.D.N.' → 'Frente Amplio'; repara mojibake conocido."""
    s = conv or ""
    s = re.sub(r"_-_O\.D\.[ND]\.?$", "", s)          # quita sufijo de órgano
    s = re.sub(r"_-_O\.D\.[ND]$", "", s)
    s = s.replace("_", " ").strip()
    # mojibake: 'De la Concertacin' → 'De la Concertación'
    fixes = {
        "De la Concertacin": "De la Concertación",
        "Union para el Cambio": "Unión para el Cambio",
    }
    return fixes.get(s, s)


def header_index(ws):
    hdr = list(next(ws.iter_rows(min_row=9, max_row=9, values_only=True)))
    return {h: i for i, h in enumerate(hdr) if h is not None}, hdr


def convert():
    rows_out = []
    odn_files = sorted(glob.glob(os.path.join(RAW, "odn", "*.xlsx")))
    odd_files = sorted(glob.glob(os.path.join(RAW, "odd", "*.xlsx")))
    stats = {"HOJA_ODN": 0, "HOJA_ODD": 0, "concert_wide": 0, "files": 0}

    for f in odn_files + odd_files:
        is_odn = os.path.dirname(f).endswith("odn")
        tiporeg = "HOJA_ODN" if is_odn else "HOJA_ODD"
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.active
        idx, hdr = header_index(ws)
        stats["files"] += 1

        # CASO ESPECIAL: De La Concertación ODN (formato ancho con col 'JOSE LUIS VERA')
        concert_wide = "JOSE LUIS VERA" in idx
        ser_key = "SERIES" if "SERIES" in idx else "SERIE"

        for r in ws.iter_rows(min_row=10, values_only=True):
            if r[idx["CONVOCATORIA"]] is None:
                continue
            lema = lema_from_convocatoria(r[idx["CONVOCATORIA"]])
            dep = dep_code(r[idx["DEPTO"]])
            crv = r[idx["CIRCUITO"]]
            ser = (r[idx[ser_key]] or "") if ser_key in idx else ""

            if concert_wide:
                votos = r[idx["JOSE LUIS VERA"]] or 0
                if votos <= 0:
                    continue
                rows_out.append([tiporeg, dep, crv, ser, lema,
                                 "JOSE LUIS VERA", "No aplica", votos])
                stats["concert_wide"] += 1
                continue

            votos = r[idx["CNT_VOTOS"]] or 0
            if votos <= 0:
                continue
            hoja = r[idx["HOJA"]]
            d1 = r[idx["PRECANDIDATO"]] if is_odn and "PRECANDIDATO" in idx else "No aplica"
            d1 = d1 if (d1 not in (None, "")) else "No aplica"
            rows_out.append([tiporeg, dep, crv, ser, lema, d1, hoja, votos])
            stats[tiporeg] += 1
        wb.close()
        print(f"  ✓ {os.path.basename(f):32s} {tiporeg}{' [wide concert]' if concert_wide else ''}")

    with open(OUT, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["TIPO_REGISTRO", "DEPARTAMENTO", "CRV", "SERIES", "LEMA",
                    "DESCRIPCIÓN_1", "DESCRIPCIÓN_2", "CANTIDAD_VOTOS"])
        w.writerows(rows_out)

    total = sum(int(r[7]) for r in rows_out)
    print(f"\n=== {OUT} ===")
    print(f"filas: {len(rows_out)} | HOJA_ODN={stats['HOJA_ODN']} (+{stats['concert_wide']} concert) "
          f"HOJA_ODD={stats['HOJA_ODD']} | total votos={total:,}")


if __name__ == "__main__":
    convert()
