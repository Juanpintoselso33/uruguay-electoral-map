#!/usr/bin/env python3
"""
Convierte los XLSX 2015 (esquema propio de la Corte, sin TIPO_REGISTRO) al formato
'desglose' que consumen build-{votes,hoja}-{local,circuito}.py (read_desglose):
    TIPO_REGISTRO,DEPARTAMENTO,CRV,SERIES,LEMA,DESCRIPCION_1,DESCRIPCION_2,CANTIDAD_VOTOS

donde DEPARTAMENTO = código de 2 letras (CODE_DIR), DESCRIPCION_1 = HOJA (verbatim,
incl. sufijo municipal '26371-B'), y TIPO_REGISTRO:
    departamentales-2015 → HOJA_ED  (→ contienda 'junta')
    municipales-2015      → HOJA_EM  (→ contienda 'municipio')

Salidas:
    data/raw/electoral/2015/desglose-departamental.csv
    data/raw/electoral/2015/desglose-municipal.csv

El catálogo (build-elecciones-2015.py) ya emite opcionId {junta,municipio}-{lema}-{hoja};
el resolver de los builders los reconstruye desde el catalogo.json. Filtro de lema = los 8
LEMA del catálogo (idéntico a build-elecciones-2015) → Σ(hoja) == votes.json por construcción.

Uso: python scripts/build-2015-desglose.py
"""
import csv, os, unicodedata
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data/raw/electoral/2015")

# slug → código de 2 letras (inverso de CODE_DIR de los builders). Solo los 6 cubiertos
# importan, pero se incluyen todos para robustez.
SLUG_CODE = {"artigas": "AR", "canelones": "CA", "cerro_largo": "CL", "colonia": "CO",
    "durazno": "DU", "florida": "FD", "flores": "FS", "lavalleja": "LA", "maldonado": "MA",
    "montevideo": "MO", "paysandu": "PA", "rio_negro": "RN", "rocha": "RO", "rivera": "RV",
    "salto": "SA", "san_jose": "SJ", "soriano": "SO", "tacuarembo": "TA", "treinta_y_tres": "TT"}
NAME_SLUG = {"ARTIGAS": "artigas", "CANELONES": "canelones", "CERRO LARGO": "cerro_largo",
    "COLONIA": "colonia", "DURAZNO": "durazno", "FLORES": "flores", "FLORIDA": "florida",
    "LAVALLEJA": "lavalleja", "MALDONADO": "maldonado", "MONTEVIDEO": "montevideo",
    "PAYSANDU": "paysandu", "RIO NEGRO": "rio_negro", "RIVERA": "rivera", "ROCHA": "rocha",
    "SALTO": "salto", "SAN JOSE": "san_jose", "SORIANO": "soriano", "TACUAREMBO": "tacuarembo",
    "TREINTA Y TRES": "treinta_y_tres"}
# mismos 8 lemas que build-elecciones-2015 (filtro EXACTO → reconciliación con votes.json)
LEMA_KEEP = {"FRENTE AMPLIO", "PARTIDO NACIONAL", "PARTIDO COLORADO", "PARTIDO DE LA CONCERTACION",
    "PARTIDO INDEPENDIENTE", "PARTIDO ECOLOGISTA RADICAL INTRANSIGENTE", "PARTIDO ASAMBLEA POPULAR",
    "PARTIDO DE LOS TRABAJADORES"}
# nombre de LEMA tal cual lo escribe el catálogo (LEMA_NOMBRE de build-elecciones-2015), para
# que el resolver (name2lema, normalizado) matchee. El crudo trae el nombre largo en MAYUS.
LEMA_CANON = {"FRENTE AMPLIO": "Frente Amplio", "PARTIDO NACIONAL": "Partido Nacional",
    "PARTIDO COLORADO": "Partido Colorado", "PARTIDO DE LA CONCERTACION": "Partido de la Concertación",
    "PARTIDO INDEPENDIENTE": "Partido Independiente",
    "PARTIDO ECOLOGISTA RADICAL INTRANSIGENTE": "Partido Ecologista Radical Intransigente",
    "PARTIDO ASAMBLEA POPULAR": "Asamblea Popular",
    "PARTIDO DE LOS TRABAJADORES": "Partido de los Trabajadores"}


def normU(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn").upper().strip()


def convert(xlsx, with_muni, tiporeg, out):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    started = False
    rows = []
    for r in ws.iter_rows(values_only=True):
        if not started:
            if r and r[0] == "ACTO":
                started = True
            continue
        if not r or r[0] is None:
            continue
        if with_muni:
            _, _, dep, _muni, crv, serie, _esc, lema, hoja, votos = r[:10]
        else:
            _, _, dep, crv, serie, _esc, lema, hoja, votos = r[:9]
        slug = NAME_SLUG.get(normU(dep))
        if not slug:
            continue
        lk = normU(lema)
        if lk not in LEMA_KEEP:
            continue
        try:
            v = int(votos or 0)
        except (ValueError, TypeError):
            continue
        if v == 0:
            continue
        rows.append({
            "TIPO_REGISTRO": tiporeg,
            "DEPARTAMENTO": SLUG_CODE[slug],
            "CRV": str(crv).strip() if crv is not None else "",
            "SERIES": (str(serie).strip().upper() if serie is not None else ""),
            "LEMA": LEMA_CANON[lk],
            "DESCRIPCION_1": (str(hoja).strip() if hoja is not None else ""),
            "DESCRIPCION_2": "",
            "CANTIDAD_VOTOS": v,
        })
    wb.close()
    with open(out, "w", encoding="utf-8", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["TIPO_REGISTRO", "DEPARTAMENTO", "CRV", "SERIES",
            "LEMA", "DESCRIPCION_1", "DESCRIPCION_2", "CANTIDAD_VOTOS"])
        wr.writeheader()
        wr.writerows(rows)
    print(f"  {os.path.basename(out)}: {len(rows)} filas (TIPO_REGISTRO={tiporeg})")
    return len(rows)


def main():
    convert(os.path.join(RAW, "deptal-2015.xlsx"), False, "HOJA_ED",
            os.path.join(RAW, "desglose-departamental.csv"))
    convert(os.path.join(RAW, "municipal-2015.xlsx"), True, "HOJA_EM",
            os.path.join(RAW, "desglose-municipal.csv"))


if __name__ == "__main__":
    main()
