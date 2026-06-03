#!/usr/bin/env python3
"""Parsea la Integración de hojas de votación → public/data/personas/personas-hoja.{eleccion}.json:
un registro por (persona × hoja × cargo). id de persona = credencial (estable cross-elección).
Uso: python scripts/build-personas-hoja.py nacionales-2024
"""
import csv, json, os, sys, unicodedata

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CODE_DEPT = {"AR":"artigas","CA":"canelones","CL":"cerro_largo","CO":"colonia","DU":"durazno",
    "FD":"florida","FS":"flores","LA":"lavalleja","MA":"maldonado","MO":"montevideo","PA":"paysandu",
    "RN":"rio_negro","RO":"rocha","RV":"rivera","SA":"salto","SJ":"san_jose","SO":"soriano",
    "TA":"tacuarembo","TT":"treinta_y_tres"}

# Valores que indican "sin dato" — cubren: vacío, "NO APLICA" (con/sin punto),
# "SIN SUBLEMA" y "SIN AGRUPACIÓN" / "SIN AGRUPACION" (con o sin tilde).
VACIO = {"", "NO APLICA", "NO APLICA.", "N/A", "SIN SUBLEMA", "SIN AGRUPACION", "SIN AGRUPACIÓN"}


def clean(s):
    return (s or "").strip()


def norm_upper(s):
    """Normaliza a mayúsculas quitando tildes (para comparar contra VACIO)."""
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper().strip()


def persona_id(serie, numero):
    return f"{clean(serie).upper()}-{clean(numero)}"


def read_rows(eleccion):
    base = os.path.join(ROOT, "data/raw/electoral", eleccion)
    # Candidatos de nombre para CSV: el full se prueba ANTES del canónico para evitar
    # que un nominas-integracion.csv truncado (dep-2025: solo Artigas) tape al completo.
    csv_candidates = [
        "integracion-de-hojas-full.csv",   # departamentales-2025 full (19 deptos) — prioritario
        "nominas-integracion.csv",
        "integracion-de-hojas.csv",
    ]
    xlsx_candidates = [
        "nominas-integracion.xlsx",
        "integracion-hojas-de-votacion.xlsx",  # internas-2024
        "integracion-de-hojas.xlsx",
    ]
    for name in csv_candidates:
        csv_path = os.path.join(base, name)
        if os.path.exists(csv_path) and os.path.getsize(csv_path) > 1024:
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    with open(csv_path, encoding=enc) as f:
                        rows = list(csv.DictReader(f))
                    if rows:
                        print(f"  leyendo {name} (enc={enc})")
                        return rows
                except UnicodeDecodeError:
                    continue
    from openpyxl import load_workbook
    for name in xlsx_candidates:
        xlsx = os.path.join(base, name)
        if os.path.exists(xlsx):
            print(f"  leyendo {name}")
            ws = load_workbook(xlsx, read_only=True).worksheets[0]
            it = ws.iter_rows(values_only=True)
            header = [str(c).strip() if c is not None else "" for c in next(it)]
            rows = [dict(zip(header, [("" if v is None else str(v)) for v in row])) for row in it]
            # Descartar fila de encabezado repetida (internas-2024 tiene header en fila 2)
            if rows and rows[0].get("Numero", "").lower() in ("numero", "número"):
                rows = rows[1:]
            return rows
    raise SystemExit(f"no hay integración para {eleccion} (corré etl:nominas-fetch)")


def main():
    if len(sys.argv) < 2:
        raise SystemExit("uso: build-personas-hoja.py <eleccion>")
    eleccion = sys.argv[1]
    rows = read_rows(eleccion)
    out = []
    for r in rows:
        cod = clean(r.get("Departamento")).upper()
        sub_raw = clean(r.get("Sublema"))
        agr_raw = clean(r.get("Agrupacion"))
        # "Candidatura" en nacionales/departamentales; "TipoHoja" en internas
        cargo_raw = clean(r.get("Candidatura") or r.get("TipoHoja") or "")
        rec = {
            "personaId": persona_id(r.get("CredencialSerie"), r.get("CredencialNumero")),
            "nombre": clean(r.get("Nombre")),
            "sexo": clean(r.get("Sexo")) or None,
            "departamento": CODE_DEPT.get(cod, cod.lower()),
            "hoja": clean(r.get("Numero")),
            "partido": clean(r.get("PartidoPolitico")),
            "agrupacion": agr_raw if norm_upper(agr_raw) not in VACIO else None,
            "sublema": sub_raw if norm_upper(sub_raw) not in VACIO else None,
            "cargo": cargo_raw,
            "ordinal": clean(r.get("Ordinal")) or None,
            "titularSuplente": clean(r.get("TitularSuplente")) or None,
        }
        out.append(rec)
    destdir = os.path.join(ROOT, "public/data/personas")
    os.makedirs(destdir, exist_ok=True)
    dest = os.path.join(destdir, f"personas-hoja.{eleccion}.json")
    with open(dest, "w", encoding="utf-8") as f:
        json.dump({"eleccion": eleccion, "registros": out}, f, ensure_ascii=False)
    from collections import Counter
    cargos = Counter(r["cargo"] for r in out)
    print(f"{eleccion}: {len(out)} registros, {len({r['personaId'] for r in out})} personas únicas")
    print(f"cargos: {dict(cargos.most_common())}")
    size_mb = os.path.getsize(dest) / 1024 / 1024
    print(f"tamaño: {size_mb:.2f} MB → {dest}")


if __name__ == "__main__":
    main()
