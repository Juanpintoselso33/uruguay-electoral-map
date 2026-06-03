#!/usr/bin/env python3
"""
build-internas-sublemas.py — inyecta el nivel SUBLEMA en las contiendas ODN y ODD del catálogo de internas.

El acordeón (OpcionAccordion) y el gate de escaleras soportan niveles intermedios. El desglose-de-votos
NO trae la columna SUBLEMA; sí la trae la "Integración hojas de votación" de la Corte (hoja → sublema,
separado por TipoHoja=ODN/ODD).

Jerarquías resultantes:
  ODD (Convención Departamental): lema → sublema → hoja               (sublema cuelga del LEMA)
  ODN (Convención Nacional):      lema → precandidato → sublema → hoja (sublema cuelga del PRECANDIDATO)

Este post-step (espeja build-hoja-local.py: Python sobre los shards ya generados) PARCHEA en sitio el
`catalogo.json` de cada depto: agrega nodos `sublema`, pone `grupoId`+`sublemaId` en cada opción agrupada,
y ajusta `niveles` (degradado a niveles más cortos + degradado:true si un depto no tiene sublemas con dato).
Preserva los lemaId/precandidatoId existentes (parchea, no regenera) y no toca los shards de votos. Idempotente.

REQUIERE que ESCALERAS declare (granularidad.ts):
  (internas, odn) = ['lema','precandidato','sublema','hoja']
  (internas, odd) = ['lema','sublema','hoja']

Fuente de integración (auto-detección por elección):
  internas-2024 → data/raw/electoral/internas-2024/integracion-hojas-de-votacion.xlsx (hoja 'Datos')
  internas-2019 → data/raw/electoral/internas-2019/integraci-n-de-las-hojas-de-votaci-n.csv
(el CSV de 2024 viene vacío del origen → se usa el XLSX).

Uso:
  python scripts/build-internas-sublemas.py internas-2024
  python scripts/build-internas-sublemas.py internas-2019
"""
import csv, json, os, re, glob, unicodedata, argparse
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# folder de depto → código de 2 letras de la Corte
DEPT_CODE = {"artigas":"AR","canelones":"CA","cerro_largo":"CL","colonia":"CO","durazno":"DU",
    "florida":"FD","flores":"FS","lavalleja":"LA","maldonado":"MA","montevideo":"MO","paysandu":"PA",
    "rio_negro":"RN","rocha":"RO","rivera":"RV","salto":"SA","san_jose":"SJ","soriano":"SO",
    "tacuarembo":"TA","treinta_y_tres":"TT"}

SIN_SUBLEMA = {"SIN SUBLEMA", "NO APLICA", "NO APLICA.", ""}

# contienda → (TipoHoja en integración, modo de padre del sublema, escalera completa)
CONTIENDAS = {
    "odn": {"tipo": "ODN", "parent": "precandidato", "niveles": ["lema", "precandidato", "sublema", "hoja"]},
    "odd": {"tipo": "ODD", "parent": "lema",         "niveles": ["lema", "sublema", "hoja"]},
}


def norm(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper().strip()


def norm_strip(s):
    """norm + saca el prefijo 'PARTIDO ' → une 'PARTIDO NACIONAL' ↔ 'Partido Nacional' ↔ 'Nacional'."""
    return re.sub(r"^PARTIDO\s+", "", norm(s))


def slug(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def integracion_source(eleccion):
    base = os.path.join(ROOT, "data/raw/electoral", eleccion)
    for name in os.listdir(base):
        low = name.lower()
        if "integ" in low and "hoja" in low and low.endswith(".csv") and os.path.getsize(os.path.join(base, name)) > 0:
            return ("csv", os.path.join(base, name))
    for name in os.listdir(base):
        low = name.lower()
        if "integ" in low and "hoja" in low and low.endswith(".xlsx"):
            return ("xlsx", os.path.join(base, name))
    raise SystemExit(f"  ✗ sin integración de hojas en {base}")


def read_integracion_rows(kind, path):
    if kind == "csv":
        enc = "utf-8-sig"
        try: open(path, encoding=enc).readline()
        except Exception: enc = "latin-1"
        with open(path, encoding=enc, errors="replace", newline="") as f:
            for row in csv.DictReader(f):
                yield row
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Datos"] if "Datos" in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(h) if h is not None else "" for h in next(it)]
        idx = {h: i for i, h in enumerate(hdr)}
        for row in it:
            yield {h: row[idx[h]] for h in hdr}


def build_sublema_maps(eleccion):
    """Devuelve {tipo: (submap, present)} para TipoHoja ∈ {ODN, ODD}.
       submap[(deptCode, normStripPartido, hojaStr)] = sublemaRaw (solo sublema real).
       present = todas las hojas de ese tipo (incl. 'Sin sublema') → distingue 'directa' de 'bug de join'."""
    kind, path = integracion_source(eleccion)
    print(f"  fuente integración: [{kind}] {os.path.relpath(path, ROOT)}")
    sub = {t: {} for t in ("ODN", "ODD")}
    pres = {t: set() for t in ("ODN", "ODD")}
    conflicts = 0
    for row in read_integracion_rows(kind, path):
        tipo = str(row.get("TipoHoja", "")).strip().upper()
        if tipo not in sub:
            continue
        dep = str(row.get("Departamento", "") or "").strip().upper()
        part = norm_strip(str(row.get("PartidoPolitico", "") or ""))
        hoja = str(row.get("Numero", "") or "").strip()
        if hoja.endswith(".0"):
            hoja = hoja[:-2]
        if not (dep and part and hoja):
            continue
        key = (dep, part, hoja)
        pres[tipo].add(key)
        s = str(row.get("Sublema", "") or "").strip()
        if norm(s) in SIN_SUBLEMA:
            continue
        if key in sub[tipo] and sub[tipo][key] != s:
            conflicts += 1
        sub[tipo][key] = s
    if conflicts:
        print(f"  ⚠ {conflicts} (dep,partido,hoja) con sublema ambiguo (gana el último)")
    return {t: (sub[t], pres[t]) for t in sub}


def patch_contienda(c, dep_code, submap, present, parent_mode, niveles_full):
    """Parchea una contienda (odn/odd) in-place. Devuelve métricas."""
    lema_key = {n["id"]: norm_strip(n.get("etiqueta", "")) for n in c["nodos"] if n.get("nivel") == "lema"}
    # idempotencia: sacar nodos sublema previos (conserva lema/precandidato)
    c["nodos"] = [n for n in c["nodos"] if n.get("nivel") != "sublema"]
    sub_nodes = {}
    n_group = n_direct = n_unmatched = 0
    for op in c.get("opciones", []):
        op.pop("grupoId", None); op.pop("sublemaId", None)  # conserva precandidatoId
        lemaId = op.get("lemaId", "")
        hoja = str(op.get("hoja", ""))
        part = lema_key.get(lemaId, "")
        sub = submap.get((dep_code, part, hoja))
        if not sub:
            if hoja and hoja != "vl" and (dep_code, part, hoja) not in present:
                n_unmatched += 1
            n_direct += 1
            continue
        parent = (op.get("precandidatoId") or lemaId) if parent_mode == "precandidato" else lemaId
        subId = f"{parent}-sl-{slug(sub)}"
        if subId not in sub_nodes:
            sub_nodes[subId] = {"id": subId, "nivel": "sublema", "etiqueta": sub, "parentId": parent}
        op["sublemaId"] = subId
        op["grupoId"] = subId
        n_group += 1
    c["nodos"].extend(sub_nodes.values())
    if sub_nodes:
        c["niveles"] = list(niveles_full)
        c.pop("degradado", None)
    else:
        # degradado: escalera sin el nivel sublema (subsecuencia) + flag
        c["niveles"] = [n for n in niveles_full if n != "sublema"]
        c["degradado"] = True
    return {"sublemas": len(sub_nodes), "agrupadas": n_group, "directas": n_direct, "sin_match": n_unmatched}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("eleccion", help="internas-2024 | internas-2019")
    a = ap.parse_args()

    print(f"=== Inyección de SUBLEMA en internas (ODN + ODD) — {a.eleccion} ===")
    maps = build_sublema_maps(a.eleccion)
    for t in ("ODN", "ODD"):
        print(f"  {t}: {len(maps[t][0])} hoja→sublema · {len(maps[t][1])} hojas en integración")

    cats = sorted(glob.glob(os.path.join(ROOT, f"public/data/{a.eleccion}/*/catalogo.json")))
    tot = defaultdict(lambda: {"sub": 0, "grp": 0, "unm": 0})
    for p in cats:
        dep = os.path.basename(os.path.dirname(p))
        code = DEPT_CODE.get(dep)
        if not code:
            print(f"  · {dep}: sin código — skip"); continue
        doc = json.load(open(p, encoding="utf-8"))
        line = [f"{dep:16s}"]
        for cont, cfg in CONTIENDAS.items():
            c = next((x for x in doc.get("contiendas", []) if x.get("contienda") == cont), None)
            if not c:
                continue
            submap, present = maps[cfg["tipo"]]
            r = patch_contienda(c, code, submap, present, cfg["parent"], cfg["niveles"])
            tot[cont]["sub"] += r["sublemas"]; tot[cont]["grp"] += r["agrupadas"]; tot[cont]["unm"] += r["sin_match"]
            warn = f"⚠{r['sin_match']}" if r["sin_match"] else ""
            deg = "(deg)" if not r["sublemas"] else ""
            line.append(f"{cont.upper()} {r['sublemas']:3d}sl/{r['agrupadas']:4d}h{warn}{deg}")
        json.dump(doc, open(p, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        with open(p, "a", encoding="utf-8") as f:
            f.write("\n")
        print("  ✓ " + "  ".join(line))
    for cont in CONTIENDAS:
        t = tot[cont]
        print(f"{cont.upper()}: {t['sub']} nodos sublema · {t['grp']} hojas agrupadas · {t['unm']} sin match.")


if __name__ == "__main__":
    main()
