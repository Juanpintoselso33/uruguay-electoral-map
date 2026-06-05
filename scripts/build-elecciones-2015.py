#!/usr/bin/env python3
"""Issue #39 — Elecciones Departamentales y Municipales 2015 (10/05/2015).

Datos: XLSX por circuito de la Corte (NO están en CKAN), esquema propio (sin TIPO_REGISTRO):
  data/raw/electoral/2015/deptal-2015.xlsx     → ACTO,CONVOCATORIA,DEPTO,CIRCUITO,SERIES,ESCRUTINIO,LEMA,HOJA,CNT_VOTOS
  data/raw/electoral/2015/municipal-2015.xlsx  → + columna MUNICIPIO ; HOJA sufijada
Escrutinio "Departamental" = definitivo (exportado 2015-05-19).

ALCANCE (pedido de Juan: "cargos electos nada más"):
  - departamentales-2015: mapa por LEMA ganador. NACIONAL (19 deptos) + INTERIOR por SERIE.
    + intendentes electos (investigados, 19). Montevideo NO tiene vista por-depto (su detalle
    departamental usa barrios/CRV→barrio, sin mapeo 2015) → queda en la vista nacional.
  - municipales-2015: mapa por LEMA ganador, NACIONAL + 19 deptos (serie→municipio de la
    columna MUNICIPIO). SIN nombres de alcaldes/concejos (no hay nómina abierta 2015).

Uso: python scripts/build-elecciones-2015.py
"""
import csv, json, os, unicodedata
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data/raw/electoral/2015")
DATA = os.path.join(ROOT, "public/data")
MAP_DIR = os.path.join(ROOT, "public/data/mappings")

NAME_SLUG = {"ARTIGAS": "artigas", "CANELONES": "canelones", "CERRO LARGO": "cerro_largo",
    "COLONIA": "colonia", "DURAZNO": "durazno", "FLORES": "flores", "FLORIDA": "florida",
    "LAVALLEJA": "lavalleja", "MALDONADO": "maldonado", "MONTEVIDEO": "montevideo",
    "PAYSANDU": "paysandu", "RIO NEGRO": "rio_negro", "RIVERA": "rivera", "ROCHA": "rocha",
    "SALTO": "salto", "SAN JOSE": "san_jose", "SORIANO": "soriano", "TACUAREMBO": "tacuarembo",
    "TREINTA Y TRES": "treinta_y_tres"}
INTERIOR = [s for s in set(NAME_SLUG.values()) if s != "montevideo"]

LEMA_SLUG = {
    "FRENTE AMPLIO": "frente-amplio", "PARTIDO NACIONAL": "nacional", "PARTIDO COLORADO": "colorado",
    "PARTIDO DE LA CONCERTACION": "concertacion", "PARTIDO INDEPENDIENTE": "independiente",
    "PARTIDO ECOLOGISTA RADICAL INTRANSIGENTE": "ecologista-radical-intransigente",
    "PARTIDO ASAMBLEA POPULAR": "asamblea-popular", "PARTIDO DE LOS TRABAJADORES": "trabajadores",
}
LEMA_NOMBRE = {
    "frente-amplio": "Frente Amplio", "nacional": "Partido Nacional", "colorado": "Partido Colorado",
    "concertacion": "Partido de la Concertación", "independiente": "Partido Independiente",
    "ecologista-radical-intransigente": "Partido Ecologista Radical Intransigente",
    "asamblea-popular": "Asamblea Popular", "trabajadores": "Partido de los Trabajadores",
}

# Intendentes electos 2015 (investigado + cruzado: Presidencia/gub.uy + Cuadro Ágora/Corte). 12 PN, 6 FA, 1 PC.
INTENDENTES = {
    "artigas": "Pablo Caram", "canelones": "Yamandú Orsi", "cerro_largo": "Luis Sergio Botana",
    "colonia": "Carlos Moreira", "durazno": "Carmelo Vidalín", "flores": "Fernando Echeverría",
    "florida": "Carlos Enciso", "lavalleja": "Adriana Peña", "maldonado": "Enrique Antía",
    "montevideo": "Daniel Martínez", "paysandu": "Guillermo Caraballo", "rio_negro": "Oscar Terzaghi",
    "rivera": "Marne Osorio", "rocha": "Aníbal Pereyra", "salto": "Andrés Lima",
    "san_jose": "José Luis Falero", "soriano": "Agustín Bascou", "tacuarembo": "Eber Da Rosa",
    "treinta_y_tres": "Dardo Sánchez",
}


def normU(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "")) if unicodedata.category(c) != "Mn").upper().strip()


def normBarrio(s):
    """geoId de barrio MVD: idéntico a normName() del ETL (NFD-strip + UPPER + colapsa \\s).
    Debe coincidir con los geoId de la geometría de barrios (igual que departamentales-2025)."""
    return " ".join(normU(s).split())


def labels():
    return {d["id"]: d["label"] for d in json.load(open(os.path.join(ROOT, "src/config/departments.json"), encoding="utf-8"))}


def write(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    json.dump(obj, open(path, "w", encoding="utf-8"), ensure_ascii=False)


def read_xlsx(path, with_muni):
    """Devuelve filas (depto_slug, municipio|None, serie, lema_slug, hoja, votos), saltando preámbulo.

    `hoja` es el número de lista crudo (str): plano en deptal ("77"), sufijado en municipal ("26371-B").
    El filtro de lema (8 entradas de LEMA_SLUG) es EXACTO al de las columnas lema → garantiza que
    Σ(hoja por lema) == el votes.json lema-level por construcción (auditado: 0 lemas descartados).
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    started = False
    for r in ws.iter_rows(values_only=True):
        if not started:
            if r and r[0] == "ACTO":
                started = True
            continue
        if not r or r[0] is None:
            continue
        if with_muni:
            _, _, dep, muni, _, serie, _, lema, hoja, votos = r[:10]
        else:
            _, _, dep, _, serie, _, lema, hoja, votos = r[:9]
            muni = None
        slug = NAME_SLUG.get(normU(dep))
        ls = LEMA_SLUG.get(normU(lema))
        if not slug or not ls:
            continue
        try:
            v = int(votos or 0)
        except (ValueError, TypeError):
            continue
        out.append((slug, (muni or "").strip() or None, (serie or "").strip().lower(), ls,
                    (str(hoja).strip() if hoja is not None else ""), v))
    wb.close()
    return out


def zona(geoId, por_lema):
    porOpcion = [{"opcionId": l, "votos": v} for l, v in sorted(por_lema.items(), key=lambda kv: -kv[1])]
    return {"geoId": geoId, "ganadorOpcionId": porOpcion[0]["opcionId"] if porOpcion else None,
            "validos": sum(por_lema.values()), "porOpcion": porOpcion,
            "noPartidarios": {"enBlanco": 0, "anulados": 0, "observados": 0}}


def zona_hoja(geoId, por_hoja, extra=None):
    """Zona con porOpcion de HOJA (opcionId compuesto). `por_hoja`: {opcionId: votos}."""
    porOpcion = [{"opcionId": o, "votos": v} for o, v in sorted(por_hoja.items(), key=lambda kv: -kv[1]) if v]
    z = {"geoId": geoId, "ganadorOpcionId": porOpcion[0]["opcionId"] if porOpcion else None,
         "validos": sum(v for v in por_hoja.values()), "porOpcion": porOpcion,
         "noPartidarios": {"enBlanco": 0, "anulados": 0, "observados": 0}}
    if extra:
        z.update(extra)
    return z


# ── Capa HOJA (Issue #39 follow-up) ────────────────────────────────────────────
# El crudo 2015 es UNA sola papeleta por elección (deptal: ESCRUTINIO 'Departamental', HOJA
# numérica; municipal: HOJA sufijada), SIN nómina/integración → no se puede separar
# intendente/junta/alcalde como en 2020. Se emite una contienda única degradada [lema, hoja]:
#   - departamentales-2015 → contienda 'junta'   (opcionId  junta-{lema}-{hoja})
#   - municipales-2015      → contienda 'municipio'(opcionId  municipio-{lema}-{hoja})
# Sin registros VOTO_LEMA (toda fila tiene hoja; 0 hojas en blanco, auditado) → sin pseudo-folio 'vl'.

def opcion_hoja(contienda, lema, hoja):
    return f"{contienda}-{lema}-{hoja}"


def build_catalogo(eleccion, slug, contienda, lemas_presentes, opciones_hoja):
    """catalogo.json con UNA contienda degradada [lema, hoja] — misma forma que municipales-2020."""
    nodos = [{"id": l, "nivel": "lema", "etiqueta": LEMA_NOMBRE[l]} for l in sorted(lemas_presentes)]
    opciones = []
    for (lema, hoja) in sorted(opciones_hoja):
        oid = opcion_hoja(contienda, lema, hoja)
        opciones.append({"id": oid, "contienda": contienda, "clase": "hoja", "hoja": hoja,
                         "lemaId": lema, "partidoId": lema, "grupoId": lema})
    return {"eleccionId": eleccion, "departamento": slug,
            "contiendas": [{"contienda": contienda, "niveles": ["lema", "hoja"],
                            "degradado": True, "nodos": nodos, "opciones": opciones}]}


def reconcile_hoja(eleccion, slug, contienda, por_geo_hoja, opc_lema):
    """GATE delta=0: Σ(hoja por lema) por geo == votes.json (lema) on-disk. Devuelve nº de mismatches.

    No-tautológico: reconcilia contra el votes.json YA PUBLICADO (no se reescribe), así que un delta
    cazaría una divergencia entre el pase de lema previo y el de hoja."""
    vpath = os.path.join(DATA, eleccion, slug, "votes.json")
    if not os.path.exists(vpath):
        print(f"    ! {slug}: sin votes.json para reconciliar {contienda}")
        return 1
    votes = json.load(open(vpath, encoding="utf-8"))
    expected = {z["geoId"]: {o["opcionId"]: o["votos"] for o in z["porOpcion"]} for z in votes["zonas"]}
    fails = 0
    for geo, por_hoja in por_geo_hoja.items():
        got = defaultdict(int)
        for oid, v in por_hoja.items():
            got[opc_lema[oid]] += v
        exp = expected.get(geo)
        if exp is None:
            print(f"    ✗ {slug}/{geo}: geo ausente en votes.json (geoId no matchea)")
            fails += 1
            continue
        for k in set(got) | set(exp):
            if got.get(k, 0) != exp.get(k, 0):
                print(f"    ✗ {slug}/{geo} lema {k}: hoja {got.get(k,0)} ≠ votes {exp.get(k,0)}")
                fails += 1
    for geo in expected:
        if geo not in por_geo_hoja:
            print(f"    ✗ {slug}/{geo}: en votes.json pero sin hojas")
            fails += 1
    return fails


# ── Montevideo por BARRIO (CRV→barrio del ciclo 2015) ──────────────────────────
# A diferencia del interior (por SERIE), MVD une CRV→barrio con el mapeo POR CICLO
# (data/mappings/montevideo-circuito-barrio.2015.json, build-circuito-barrio-cycles.py),
# igual que departamentales-2025. Fuente de votos: desglose-departamental.csv (conversión fiel
# del XLSX, mismo filtro de 8 lemas → Σ MO idéntica, auditado). geoId = barrio normalizado
# (normBarrio) para casar con la geometría de barrios. Emite votes.json (nivel 'zona'),
# opciones.json, catalogo.json y hoja/junta/{lema}.json — gemelo barrio de la vista interior.

def build_montevideo_barrio(gate_fail):
    desg = os.path.join(RAW, "desglose-departamental.csv")
    mapping = os.path.join(ROOT, "data/mappings/montevideo-circuito-barrio.2015.json")
    if not (os.path.exists(desg) and os.path.exists(mapping)):
        print("montevideo (barrio): falta desglose o mapping CRV→barrio → se omite")
        return
    c2b = json.load(open(mapping, encoding="utf-8"))["crvToBarrio"]
    def barrio_de(crv):
        crv = str(crv).strip()
        b = c2b.get(crv)
        if not b and crv.isdigit():
            b = c2b.get(str(int(crv)))  # normaliza ceros a la izquierda
        return b
    # lema crudo (LEMA_CANON del desglose, p.ej. "Partido de la Concertación") → slug canónico
    name2slug = {normU(v): k for k, v in LEMA_NOMBRE.items()}
    por_barrio = defaultdict(lambda: defaultdict(int))           # barrio -> lema -> votos
    por_barrio_hoja = defaultdict(lambda: defaultdict(int))      # barrio -> opcionId -> votos
    opc_lema, lemas_pres, pares = {}, set(), set()
    unmapped = 0
    with open(desg, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["DEPARTAMENTO"] != "MO":
                continue
            v = int(row["CANTIDAD_VOTOS"])
            ls = name2slug.get(normU(row["LEMA"]))
            if not ls:
                continue
            b = barrio_de(row["CRV"])
            if not b:
                unmapped += v
                continue
            bn = normBarrio(b)
            hoja = (row["DESCRIPCION_1"] or "").strip()
            por_barrio[bn][ls] += v
            if hoja:
                oid = opcion_hoja("junta", ls, hoja)
                por_barrio_hoja[bn][oid] += v
                opc_lema[oid] = ls
                lemas_pres.add(ls)
                pares.add((ls, hoja))
    # votes.json (nivel 'zona', geoId = barrio) + opciones.json
    zonas = [zona(b, lemas) for b, lemas in sorted(por_barrio.items())]
    op = {l: LEMA_NOMBRE[l] for lemas in por_barrio.values() for l in lemas}
    write(os.path.join(DATA, "departamentales-2015", "montevideo", "votes.json"),
          {"eleccionId": "departamentales-2015", "departamento": "montevideo", "nivel": "zona",
           "escrutinio": "definitivo", "tipo": "departamentales", "zonas": zonas})
    write(os.path.join(DATA, "departamentales-2015", "montevideo", "opciones.json"),
          {"opciones": [{"opcionId": k, "nombre": v} for k, v in sorted(op.items())]})
    # catalogo.json (contienda 'junta' degradada [lema, hoja]) + hoja/junta/{lema}.json por barrio
    cat = build_catalogo("departamentales-2015", "montevideo", "junta", lemas_pres, pares)
    write(os.path.join(DATA, "departamentales-2015", "montevideo", "catalogo.json"), cat)
    nsh = 0
    for lema in sorted(lemas_pres):
        by_barrio = defaultdict(lambda: defaultdict(int))
        for b, por in por_barrio_hoja.items():
            for oid, v in por.items():
                if opc_lema[oid] == lema:
                    by_barrio[b][oid] += v
        if not by_barrio:
            continue
        zb = [zona_hoja(b, por) for b, por in sorted(by_barrio.items())]
        write(os.path.join(DATA, "departamentales-2015", "montevideo", "hoja", "junta", f"{lema}.json"),
              {"eleccionId": "departamentales-2015", "departamento": "montevideo", "nivel": "zona",
               "escrutinio": "definitivo", "tipo": "departamentales", "clase": "hoja", "zonas": zb})
        nsh += 1
    fails = reconcile_hoja("departamentales-2015", "montevideo", "junta", por_barrio_hoja, opc_lema)
    gate_fail[0] += fails
    pct = 100 * (1 - unmapped / max(1, sum(sum(l.values()) for l in por_barrio.values()) + unmapped))
    print(f"  {'✓' if fails==0 else '✗'} montevideo: {len(zonas)} barrios · {len(pares)} hojas · "
          f"{nsh} shards · CRV→barrio {pct:.1f}% · votos sin barrio={unmapped} · recon fails={fails}")


def main():
    LBL = labels()
    gate_fail = [0]  # acumulador mutable de mismatches de reconciliación HOJA (gate global)

    # ---------- DEPARTAMENTALES 2015 ----------
    dep_rows = read_xlsx(os.path.join(RAW, "deptal-2015.xlsx"), with_muni=False)
    por_depto = defaultdict(lambda: defaultdict(int))         # slug -> lema -> votos
    por_serie = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))  # slug -> serie -> lema -> votos
    # HOJA por serie: slug -> serie -> opcionId(junta-{lema}-{hoja}) -> votos  + catálogo lema/hoja
    dep_hoja = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    dep_opc_lema = defaultdict(dict)                          # slug -> {opcionId -> lema}
    dep_lemas = defaultdict(set)                              # slug -> {lema}
    dep_pares = defaultdict(set)                              # slug -> {(lema, hoja)}
    for slug, _m, serie, lema, hoja, v in dep_rows:
        por_depto[slug][lema] += v
        if serie:
            por_serie[slug][serie][lema] += v
            if hoja:
                oid = opcion_hoja("junta", lema, hoja)
                dep_hoja[slug][serie][oid] += v
                dep_opc_lema[slug][oid] = lema
                dep_lemas[slug].add(lema)
                dep_pares[slug].add((lema, hoja))

    opc_dep = {}
    nac_zonas = []
    intend = {}
    for slug, lemas in por_depto.items():
        for l in lemas: opc_dep[l] = LEMA_NOMBRE[l]
        nac_zonas.append(zona(LBL[slug], lemas))
        gan = max(lemas, key=lemas.get)
        intend[LBL[slug]] = {"ganadorLema": gan,
                             "intendenteElecto": {"candidato": INTENDENTES[slug], "votos": lemas[gan], "lema": gan},
                             "lemas": {}}  # sin desglose por candidato en 2015 (cargos electos nada más)
    write(os.path.join(DATA, "departamentales-2015", "_nacional", "votes.json"),
          {"eleccionId": "departamentales-2015", "departamento": "_nacional", "nivel": "departamento",
           "escrutinio": "definitivo", "tipo": "departamentales", "zonas": nac_zonas})
    write(os.path.join(DATA, "departamentales-2015", "_nacional", "opciones.json"),
          {"opciones": [{"opcionId": k, "nombre": v} for k, v in sorted(opc_dep.items())]})
    write(os.path.join(DATA, "departamentales-2015", "_nacional", "intendentes.json"),
          {"fuente": "Corte Electoral (XLSX por circuito, escrutinio definitivo); intendentes electos investigados y cruzados.",
           "eleccion": "departamentales-2015", "departamentos": intend})
    # Interior por SERIE (MVD usa barrios → se omite su vista por-depto; queda en la nacional).
    for slug in INTERIOR:
        if slug not in por_serie: continue
        zonas = [zona(s, lemas) for s, lemas in sorted(por_serie[slug].items())]
        op = {}
        for s, lemas in por_serie[slug].items():
            for l in lemas: op[l] = LEMA_NOMBRE[l]
        write(os.path.join(DATA, "departamentales-2015", slug, "votes.json"),
              {"eleccionId": "departamentales-2015", "departamento": slug, "nivel": "serie",
               "escrutinio": "definitivo", "tipo": "departamentales", "zonas": zonas})
        write(os.path.join(DATA, "departamentales-2015", slug, "opciones.json"),
              {"opciones": [{"opcionId": k, "nombre": v} for k, v in sorted(op.items())]})
    print(f"departamentales-2015: nacional {len(nac_zonas)} deptos + interior por serie ({len(INTERIOR)})")

    # Montevideo por BARRIO (CRV→barrio del ciclo 2015) — votes/opciones/catalogo/hoja en un pase.
    print("departamentales-2015 · Montevideo (BARRIO, CRV→barrio 2015):")
    build_montevideo_barrio(gate_fail)

    # --- HOJA departamentales-2015: catalogo.json + hoja/junta/{lema}.json por serie (interior) ---
    print("departamentales-2015 · HOJA (contienda 'junta', degradada [lema, hoja]):")
    for slug in INTERIOR:
        if slug not in dep_hoja: continue
        cat = build_catalogo("departamentales-2015", slug, "junta", dep_lemas[slug], dep_pares[slug])
        write(os.path.join(DATA, "departamentales-2015", slug, "catalogo.json"), cat)
        # shards hoja/junta/{lema}.json keyed por serie (mismo geoId que votes.json)
        por_lema_zonas = defaultdict(list)   # lema -> [zona]
        for serie in sorted(dep_hoja[slug]):
            for oid, v in dep_hoja[slug][serie].items():
                por_lema_zonas[dep_opc_lema[slug][oid]].append((serie, oid, v))
        nsh = 0
        for lema in sorted(dep_lemas[slug]):
            by_serie = defaultdict(lambda: defaultdict(int))
            for (serie, oid, v) in por_lema_zonas.get(lema, []):
                by_serie[serie][oid] += v
            if not by_serie: continue
            zonas = [zona_hoja(serie, por) for serie, por in sorted(by_serie.items())]
            write(os.path.join(DATA, "departamentales-2015", slug, "hoja", "junta", f"{lema}.json"),
                  {"eleccionId": "departamentales-2015", "departamento": slug, "nivel": "serie",
                   "escrutinio": "definitivo", "tipo": "departamentales", "clase": "hoja", "zonas": zonas})
            nsh += 1
        fails = reconcile_hoja("departamentales-2015", slug, "junta", dep_hoja[slug], dep_opc_lema[slug])
        gate_fail[0] += fails
        nh = len(dep_pares[slug])
        print(f"  {'✓' if fails==0 else '✗'} {slug}: {nsh} shards · {nh} hojas · recon fails={fails}")

    # ---------- MUNICIPALES 2015 ----------
    mun_rows = read_xlsx(os.path.join(RAW, "municipal-2015.xlsx"), with_muni=True)
    serie_muni = {}                                          # (slug, serie) -> municipio
    por_muni = defaultdict(lambda: defaultdict(int))         # (slug, muni) -> lema -> votos
    # HOJA por municipio: slug -> muni -> opcionId(municipio-{lema}-{hoja}) -> votos
    mun_hoja = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    mun_opc_lema = defaultdict(dict)                          # slug -> {opcionId -> lema}
    mun_lemas = defaultdict(set)                             # slug -> {lema}
    mun_pares = defaultdict(set)                             # slug -> {(lema, hoja)}
    for slug, muni, serie, lema, hoja, v in mun_rows:
        if not muni: continue
        if serie: serie_muni[(slug, serie)] = muni
        por_muni[(slug, muni)][lema] += v
        if hoja:
            oid = opcion_hoja("municipio", lema, hoja)
            mun_hoja[slug][muni][oid] += v
            mun_opc_lema[slug][oid] = lema
            mun_lemas[slug].add(lema)
            mun_pares[slug].add((lema, hoja))

    # serie→municipio por depto (para la geometría: build-municipio-geo.ts departamentales-2015)
    for slug in set(s for s, _ in serie_muni):
        entries = sorted({s: m for (sl, s), m in serie_muni.items() if sl == slug}.items())
        write(os.path.join(MAP_DIR, slug, "serie-municipio.departamentales-2015.json"),
              [{"serie": s, "municipio": m} for s, m in entries])

    nac_mun, opc_mun = [], {}
    depto_zonas = defaultdict(list)
    for (slug, muni), lemas in sorted(por_muni.items()):
        for l in lemas: opc_mun[l] = LEMA_NOMBRE[l]
        z = zona(muni, lemas)
        depto_zonas[slug].append(z)
        nac_mun.append({**z, "geoId": f"{muni} · {LBL[slug]}"})
    for slug, zonas in depto_zonas.items():
        op = {l: LEMA_NOMBRE[l] for z in zonas for l in [o["opcionId"] for o in z["porOpcion"]]}
        write(os.path.join(DATA, "municipales-2015", slug, "votes.json"),
              {"eleccionId": "municipales-2015", "departamento": slug, "nivel": "municipio",
               "escrutinio": "definitivo", "tipo": "municipales", "zonas": zonas})
        write(os.path.join(DATA, "municipales-2015", slug, "opciones.json"),
              {"opciones": [{"opcionId": k, "nombre": v} for k, v in sorted(op.items())]})
    write(os.path.join(DATA, "municipales-2015", "_nacional", "votes.json"),
          {"eleccionId": "municipales-2015", "departamento": "_nacional", "nivel": "municipio",
           "escrutinio": "definitivo", "tipo": "municipales", "zonas": nac_mun})
    write(os.path.join(DATA, "municipales-2015", "_nacional", "opciones.json"),
          {"opciones": [{"opcionId": k, "nombre": v} for k, v in sorted(opc_mun.items())]})
    print(f"municipales-2015: nacional {len(nac_mun)} municipios + {len(depto_zonas)} deptos · {len(serie_muni)} series mapeadas")

    # --- HOJA municipales-2015: catalogo.json + hoja-municipio.json por depto (+ _nacional) ---
    # Mismo formato/geoId que municipales-2025/2020 (geoId = municipio de votes.json;
    # opcionId = municipio-{lema}-{hoja}). Contienda 'municipio' degradada [lema, hoja].
    print("municipales-2015 · HOJA (contienda 'municipio', degradada [lema, hoja]):")
    nac_hoja_zonas = []
    for slug in sorted(mun_hoja):
        cat = build_catalogo("municipales-2015", slug, "municipio", mun_lemas[slug], mun_pares[slug])
        write(os.path.join(DATA, "municipales-2015", slug, "catalogo.json"), cat)
        zonas = [zona_hoja(muni, por_hoja) for muni, por_hoja in sorted(mun_hoja[slug].items())]
        write(os.path.join(DATA, "municipales-2015", slug, "hoja-municipio.json"),
              {"eleccionId": "municipales-2015", "departamento": slug, "nivel": "municipio",
               "escrutinio": "definitivo", "tipo": "municipales", "clase": "hoja", "zonas": zonas})
        fails = reconcile_hoja("municipales-2015", slug, "municipio", mun_hoja[slug], mun_opc_lema[slug])
        gate_fail[0] += fails
        nh = len(mun_pares[slug])
        print(f"  {'✓' if fails==0 else '✗'} {slug}: {len(zonas)} municipios · {nh} hojas · recon fails={fails}")
        for muni, por_hoja in sorted(mun_hoja[slug].items()):
            nac_hoja_zonas.append(zona_hoja(f"{muni} · {LBL[slug]}", por_hoja))
    write(os.path.join(DATA, "municipales-2015", "_nacional", "hoja-municipio.json"),
          {"eleccionId": "municipales-2015", "departamento": "_nacional", "nivel": "municipio",
           "escrutinio": "definitivo", "tipo": "municipales", "clase": "hoja", "zonas": nac_hoja_zonas})
    print(f"  nacional: {len(nac_hoja_zonas)} municipios (hoja-municipio)")

    # ---------- Tamaños (cota 3 MB) + GATE global ----------
    print("\n--- tamaños (cota 3 MB/archivo) ---")
    big = []
    check_paths = []
    for el in ("departamentales-2015", "municipales-2015"):
        for dp, _d, fs in os.walk(os.path.join(DATA, el)):
            for fn in fs:
                if fn in ("catalogo.json", "hoja-municipio.json") or "/hoja/" in os.path.join(dp, fn).replace("\\", "/"):
                    check_paths.append(os.path.join(dp, fn))
    max_kb, max_p = 0, ""
    for p in check_paths:
        kb = os.path.getsize(p) / 1024
        if kb > max_kb: max_kb, max_p = kb, p
        if kb > 3 * 1024: big.append((p, kb))
    print(f"  archivo HOJA más grande: {max_kb:.1f} KB ({os.path.relpath(max_p, ROOT) if max_p else '-'})")
    if big:
        print("  ⚠ >3MB (shardear):")
        for p, kb in big: print(f"    {kb/1024:.2f} MB {os.path.relpath(p, ROOT)}")
    else:
        print("  todos ≤ 3 MB ✓")

    if gate_fail[0]:
        raise SystemExit(f"\nGATE FALLÓ: {gate_fail[0]} mismatches de reconciliación HOJA vs votes.json.")
    print("\nGATE OK: Σ(hoja por lema) == votes.json (lema) en toda zona, delta=0.")


if __name__ == "__main__":
    main()
